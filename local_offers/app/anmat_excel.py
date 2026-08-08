from __future__ import annotations

import csv
import io
import json
import logging
import re
import unicodedata
from datetime import datetime, timezone
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any
from urllib.parse import urljoin

import httpx
import xlrd
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from . import db
from .config import Settings

LOGGER = logging.getLogger(__name__)
ROOT = Path('/data/anmat')
DATASET = ROOT / 'lialg-current.bin'
META_KEY = 'anmat_excel_v1'


def _norm(v: Any) -> str:
    s = unicodedata.normalize('NFKD', str(v or '').lower().strip())
    s = ''.join(ch for ch in s if not unicodedata.combining(ch))
    return ' '.join(re.sub(r'[^a-z0-9]+', ' ', s).split())


def _header_score(row: list[Any]) -> int:
    text = ' '.join(_norm(x) for x in row)
    return sum(token in text for token in ('marca', 'denominacion', 'estado', 'rnpa'))


def _matrix_to_rows(matrix: list[list[Any]]) -> list[dict[str, str]]:
    if not matrix:
        return []
    hi = max(range(min(20, len(matrix))), key=lambda i: _header_score(matrix[i]))
    if _header_score(matrix[hi]) < 2:
        raise RuntimeError('ANMAT Excel: encabezados no reconocidos.')
    headers = [_norm(v) or f'c{i}' for i, v in enumerate(matrix[hi])]
    out = []
    for values in matrix[hi + 1:]:
        if not any(str(v or '').strip() for v in values):
            continue
        row = {headers[i]: str(values[i] or '').strip() for i in range(min(len(headers), len(values)))}
        row['_text'] = ' | '.join(str(v or '').strip() for v in values if str(v or '').strip())
        out.append(row)
    return out


def parse_export(data: bytes) -> tuple[list[dict[str, str]], str]:
    if data.startswith(b'PK'):
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        best = []
        for ws in wb.worksheets:
            try:
                rows = _matrix_to_rows([list(r) for r in ws.iter_rows(values_only=True)])
            except RuntimeError:
                continue
            if len(rows) > len(best):
                best = rows
        if best:
            return best, 'xlsx'
    if data.startswith(bytes.fromhex('D0CF11E0')):
        book = xlrd.open_workbook(file_contents=data)
        best = []
        for sheet in book.sheets():
            try:
                rows = _matrix_to_rows([sheet.row_values(i) for i in range(sheet.nrows)])
            except RuntimeError:
                continue
            if len(rows) > len(best):
                best = rows
        if best:
            return best, 'xls'
    text = data.decode('utf-8-sig', errors='ignore')
    try:
        dialect = csv.Sniffer().sniff(text[:4000], delimiters=',;\t')
        rows = _matrix_to_rows(list(csv.reader(io.StringIO(text), dialect)))
        if rows:
            return rows, 'csv'
    except Exception:
        pass
    raise RuntimeError('ANMAT: la exportación no devolvió un Excel/CSV reconocible.')


def _hidden(form) -> dict[str, str]:
    out = {}
    if form:
        for tag in form.find_all('input'):
            if tag.get('name') and str(tag.get('type') or '').lower() == 'hidden':
                out[str(tag['name'])] = str(tag.get('value') or '')
    return out


def export_candidates(html: str, page_url: str) -> list[tuple[str, str, dict[str, str]]]:
    soup = BeautifulSoup(html, 'html.parser')
    out = []
    for tag in soup.find_all(['a', 'button', 'input']):
        label = _norm(tag.get_text(' ', strip=True) if tag.name != 'input' else tag.get('value'))
        hint = _norm(f"{tag.get('href') or ''} {tag.get('formaction') or ''}")
        if 'excel' not in label and 'export' not in label and 'excel' not in hint and 'export' not in hint:
            continue
        form = tag.find_parent('form')
        payload = _hidden(form)
        if tag.get('name'):
            payload[str(tag['name'])] = str(tag.get('value') or tag.get_text(strip=True) or 'Exportar a Excel')
        action = tag.get('formaction') or tag.get('href') or (form.get('action') if form else '') or page_url
        method = str(tag.get('formmethod') or (form.get('method') if form else 'get') or 'get').lower()
        out.append((method, urljoin(page_url, str(action)), payload))
    for href in re.findall(r'[\"\']([^\"\']*(?:excel|export)[^\"\']*)[\"\']', html, flags=re.I):
        if not href.startswith(('javascript:', '#')):
            out.append(('get', urljoin(page_url, href), {}))
    base = urljoin(page_url, '/')
    for suffix in ('Home/ExportarExcel', 'Home/ExportExcel', 'Home/Exportar', 'Home/Export'):
        out.append(('get', urljoin(base, suffix), {}))
    seen, unique = set(), []
    for item in out:
        key = (item[0], item[1], json.dumps(item[2], sort_keys=True))
        if key not in seen:
            seen.add(key); unique.append(item)
    return unique


def _looks_excel(r: httpx.Response) -> bool:
    ctype = r.headers.get('content-type', '').lower()
    dispo = r.headers.get('content-disposition', '').lower()
    return r.content.startswith((b'PK', bytes.fromhex('D0CF11E0'))) or 'excel' in ctype or 'spreadsheet' in ctype or '.xls' in dispo or 'text/csv' in ctype


async def _download(settings: Settings) -> tuple[bytes, str]:
    headers = {'User-Agent': 'HA-Local-Offers/0.3.2 (+Home Assistant App)'}
    async with httpx.AsyncClient(timeout=settings.anmat_timeout_seconds, follow_redirects=True, headers=headers) as client:
        home = await client.get(settings.anmat_url); home.raise_for_status()
        errors = []
        for method, url, payload in export_candidates(home.text, str(home.url)):
            try:
                r = await (client.post(url, data=payload) if method == 'post' else client.get(url, params=payload))
                if r.status_code >= 400 or not _looks_excel(r):
                    continue
                rows, kind = parse_export(r.content)
                if rows:
                    return r.content, kind
            except Exception as exc:
                errors.append(f'{url}: {exc}')
        raise RuntimeError('No se pudo descargar Exportar a Excel de ANMAT. ' + ' | '.join(errors[-3:]))


def _meta() -> dict[str, Any]:
    try:
        return json.loads(db.get_state(META_KEY) or '{}')
    except Exception:
        return {}


def _age_hours(value: str | None) -> float | None:
    if not value:
        return None
    try:
        dt = datetime.fromisoformat(value.replace('Z', '+00:00'))
        if dt.tzinfo is None: dt = dt.replace(tzinfo=timezone.utc)
        return (datetime.now(timezone.utc) - dt.astimezone(timezone.utc)).total_seconds() / 3600
    except Exception:
        return None


async def get_dataset(settings: Settings, force: bool = False) -> dict[str, Any]:
    if not settings.anmat_enabled:
        return {'ok': False, 'rows': [], 'errors': []}
    meta = _meta(); age = _age_hours(meta.get('downloaded_at'))
    if not force and DATASET.exists() and age is not None and age <= settings.anmat_refresh_hours:
        rows, kind = parse_export(DATASET.read_bytes())
        return {'ok': True, 'rows': rows, 'kind': kind, 'cached': True, 'age_hours': round(age, 1), 'errors': []}
    try:
        data, kind = await _download(settings)
        rows, _ = parse_export(data)
        ROOT.mkdir(parents=True, exist_ok=True)
        tmp = DATASET.with_suffix('.tmp'); tmp.write_bytes(data); tmp.replace(DATASET)
        now = datetime.now(timezone.utc).isoformat()
        db.set_state(META_KEY, json.dumps({'downloaded_at': now, 'kind': kind, 'rows': len(rows)}))
        return {'ok': True, 'rows': rows, 'kind': kind, 'cached': False, 'age_hours': 0.0, 'errors': []}
    except Exception as exc:
        LOGGER.warning('No se pudo refrescar Excel ANMAT: %s', exc)
        if DATASET.exists() and age is not None and age <= settings.anmat_cache_days * 24:
            rows, kind = parse_export(DATASET.read_bytes())
            return {'ok': True, 'rows': rows, 'kind': kind, 'cached': True, 'stale_fallback': True, 'age_hours': round(age, 1), 'errors': [str(exc)]}
        return {'ok': False, 'rows': [], 'errors': [str(exc)]}


def _is_vigente(row: dict[str, str]) -> bool:
    values = [v for k, v in row.items() if 'estado' in k]
    return bool(values) and any(_norm(v) == 'vigente' for v in values)


def _row_brand(row: dict[str, str]) -> str:
    return _norm(' '.join(v for k, v in row.items() if 'marca' in k or 'fantasia' in k))


def _row_text(row: dict[str, str]) -> str:
    vals = [v for k, v in row.items() if any(t in k for t in ('marca', 'fantasia', 'denominacion', 'nombre'))]
    return _norm(' '.join(vals) if vals else row.get('_text'))


def _score(product: dict[str, Any], row: dict[str, str]) -> float:
    if not _is_vigente(row): return 0.0
    p = _norm(' '.join(str(product.get(k) or '') for k in ('brand', 'name', 'variant')))
    r = _row_text(row)
    if not p or not r: return 0.0
    brand, rb = _norm(product.get('brand')), _row_brand(row)
    if brand and rb:
        br = SequenceMatcher(None, brand, rb).ratio()
        if br < .72 and brand not in rb and rb not in brand and not (set(brand.split()) & set(rb.split())):
            return 0.0
    pt, rt = set(p.split()), set(r.split())
    seq = SequenceMatcher(None, p, r).ratio()
    jac = len(pt & rt) / max(1, len(pt | rt))
    cont = len(pt & rt) / max(1, len(pt))
    return min(1.0, seq * .35 + jac * .30 + cont * .35)


async def match_products_anmat_excel(products: list[dict[str, Any]], settings: Settings) -> dict[str, Any]:
    dataset = await get_dataset(settings)
    rows = dataset.get('rows') or []
    matches = {}
    for item in products:
        brand = _norm(item.get('brand'))
        if not brand: continue
        bt = set(brand.split())
        candidates = [r for r in rows if _is_vigente(r) and _row_brand(r) and (brand in _row_brand(r) or _row_brand(r) in brand or bt & set(_row_brand(r).split()))]
        scored = sorted(((r, _score(item, r)) for r in candidates), key=lambda x: x[1], reverse=True)
        if not scored: continue
        best, score = scored[0]; second = scored[1][1] if len(scored) > 1 else 0.0
        if score < settings.anmat_match_threshold or second >= score - .035: continue
        matches[int(item['id'])] = {'score': round(score, 3), 'state': 'Vigente', 'record': best.get('_text'), 'url': settings.anmat_url, 'dataset_downloaded_at': _meta().get('downloaded_at')}
    return {'matches': matches, 'dataset_rows': len(rows), 'dataset_kind': dataset.get('kind'), 'dataset_cached': bool(dataset.get('cached')), 'dataset_age_hours': dataset.get('age_hours'), 'errors': dataset.get('errors') or []}
